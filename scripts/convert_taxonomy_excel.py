"""把学校文件归类 Excel 转成项目内分类体系 JSON 配置。"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def main() -> None:
    """读取命令行参数并执行一次 Excel 到 JSON 的转换。"""

    parser = argparse.ArgumentParser(description="Convert taxonomy Excel to File Agent JSON config.")
    parser.add_argument("--file", required=True, help="源 Excel 文件路径。")
    parser.add_argument("--sheet", default="Sheet2", help="要读取的工作表名称，默认 Sheet2。")
    parser.add_argument("--output", required=True, help="输出 JSON 配置路径。")
    parser.add_argument("--key", default="school_file_classification", help="分类体系 key。")
    parser.add_argument("--name", default="学校文件归类表", help="分类体系名称。")
    parser.add_argument("--version", default="2026-06", help="分类体系版本。")
    args = parser.parse_args()

    payload = convert_excel_to_taxonomy(
        file_path=Path(args.file),
        sheet_name=args.sheet,
        key=args.key,
        name=args.name,
        version=args.version,
    )
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def convert_excel_to_taxonomy(
    *,
    file_path: Path,
    sheet_name: str,
    key: str,
    name: str,
    version: str,
) -> dict[str, Any]:
    """把带空白继承单元格的分类表转换为树状 JSON。"""

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    roots: list[dict[str, Any]] = []
    current_root: dict[str, Any] | None = None

    for row in worksheet.iter_rows(values_only=True):
        values = [_clean_cell(value) for value in row]
        if not any(values):
            continue
        first_value = values[0]
        if first_value:
            current_root = _node(first_value)
            roots.append(current_root)
        if current_root is None:
            continue

        branch_values = values[1:] if first_value else values[1:]
        branch_values = [value for value in branch_values if value]
        if branch_values:
            current_root["children"].append(_branch(branch_values))

    return {
        "key": key,
        "name": name,
        "version": version,
        "source": file_path.name,
        "categories": roots,
    }


def _branch(values: list[str]) -> dict[str, Any]:
    """把一行中的二级和后续分类转成父子结构。"""

    parent = _node(values[0])
    parent["children"] = [_node(value) for value in values[1:]]
    return parent


def _node(name: str) -> dict[str, Any]:
    """创建分类节点字典。"""

    return {"name": name, "children": []}


def _clean_cell(value: object) -> str:
    """规范化 Excel 单元格文本，过滤空值。"""

    return str(value).strip() if value is not None else ""


if __name__ == "__main__":
    main()
