"""从 XLS/XLSX/XLSM/CSV/TSV 原件构建受控工作簿 Profile。"""

from __future__ import annotations

import csv
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Sequence

import openpyxl

from .schemas import ColumnProfile, ColumnType, SheetProfile, WorkbookProfile


MAX_PROFILE_SAMPLE_ROWS = 100
MAX_SAMPLE_VALUES_PER_COLUMN = 5
MAX_HEADER_SCAN_ROWS = 20
SUPPORTED_SPREADSHEET_SUFFIXES = {".xls", ".xlsx", ".xlsm", ".csv", ".tsv"}

_HEADER_MARKERS = (
    "序号",
    "编号",
    "姓名",
    "申请人",
    "申报人",
    "单位",
    "部门",
    "学科",
    "系所",
    "名称",
    "类型",
    "类别",
    "数量",
    "人数",
    "金额",
    "日期",
    "时间",
    "计划",
    "要求",
    "联系方式",
    "备注",
)


def profile_workbook(
    *,
    document_id: str,
    filename: str,
    file_path: Path,
) -> WorkbookProfile:
    """读取原始工作簿结构；不读取或修改数据库，也不修改文件。"""
    suffix = file_path.suffix.lower()

    if suffix not in SUPPORTED_SPREADSHEET_SUFFIXES:
        raise ValueError("当前仅支持 .xls、.xlsx、.xlsm、.csv 和 .tsv 文件。")

    if suffix in {".csv", ".tsv"}:
        sheets = [_profile_delimited_text(file_path=file_path, suffix=suffix)]
    elif suffix == ".xls":
        raise ValueError("旧版 XLS 必须先由 Tool handler 解析为持久化 XLSX 派生件。")
    else:
        sheets = _profile_excel(file_path=file_path)

    if not sheets:
        raise ValueError("工作簿中没有可分析的工作表。")

    return WorkbookProfile(
        document_id=document_id,
        filename=filename,
        sheets=sheets,
    )


def _profile_excel(*, file_path: Path) -> list[SheetProfile]:
    workbook = openpyxl.load_workbook(
        filename=file_path,
        # 多级表头需要读取 merged_cells；只读 Worksheet 不提供完整合并区域信息。
        # 这里只构建受控 Profile，不修改或保存工作簿。
        read_only=False,
        data_only=True,
    )

    try:
        sheets: list[SheetProfile] = []

        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            header_start_row, header_row = detect_header_span(worksheet)
            headers = read_headers(
                worksheet,
                header_row,
                header_start_row=header_start_row,
            )

            if not headers:
                continue

            sheet_id = f"sheet_{sheet_index}"
            columns = build_column_profiles(
                row_iterable=worksheet.iter_rows(
                    min_row=header_row + 1,
                    values_only=True,
                ),
                sheet_id=sheet_id,
                headers=headers,
            )

            sheets.append(
                SheetProfile(
                    sheet_id=sheet_id,
                    sheet_name=worksheet.title or f"Sheet{sheet_index}",
                    header_row=header_row,
                    row_count=_count_nonempty_rows(
                        worksheet.iter_rows(
                            min_row=header_row + 1,
                            values_only=True,
                        )
                    ),
                    columns=columns,
                )
            )

        return sheets
    finally:
        workbook.close()


def _profile_delimited_text(*, file_path: Path, suffix: str) -> SheetProfile:
    """读取 CSV/TSV 的表头和列 Profile；只采样文本，不修改文件。"""

    rows = _read_delimited_rows(file_path=file_path, suffix=suffix)

    if not rows:
        raise ValueError("表格文本文件为空，无法识别表头。")

    header_row_zero_based = _first_non_empty_row_index(rows)
    header_row = header_row_zero_based + 1
    headers = _read_headers_from_values(rows[header_row_zero_based])

    if not headers:
        raise ValueError("CSV 文件未识别到有效表头。")

    columns = build_column_profiles(
        row_iterable=rows[header_row_zero_based + 1 :],
        sheet_id="sheet_1",
        headers=headers,
    )

    return SheetProfile(
        sheet_id="sheet_1",
        sheet_name="TSV" if suffix == ".tsv" else "CSV",
        header_row=header_row,
        row_count=_count_nonempty_rows(rows[header_row_zero_based + 1 :]),
        columns=columns,
    )


def detect_header_row(worksheet: Any) -> int:
    """返回数据开始前的最后一行表头，兼容单行与合并多级表头。"""

    return detect_header_span(worksheet)[1]


def detect_header_span(worksheet: Any) -> tuple[int, int]:
    """识别 Excel 表头起止行。

    标题、填报单位和日期通常位于真正表头之前，不能再把第一个非空行固定当作表头。
    候选行只根据真实单元格内容评分；确定起始行后，再使用工作簿声明的纵向合并区域扩展
    多级表头范围。无法识别业务表头时仍回退到第一个非空行，保持简单表格兼容性。
    """

    max_row = min(int(getattr(worksheet, "max_row", 1) or 1), MAX_HEADER_SCAN_ROWS)
    rows = list(
        worksheet.iter_rows(
            min_row=1,
            max_row=max_row,
            values_only=True,
        )
    )
    first_non_empty = _first_non_empty_row_index(rows) + 1

    best_row = first_non_empty
    best_score = 0
    for row_number, row in enumerate(rows, start=1):
        score = _header_candidate_score(row)
        if score > best_score:
            best_row = row_number
            best_score = score

    if best_score == 0:
        return first_non_empty, first_non_empty

    header_end = best_row
    for merged_range in getattr(worksheet.merged_cells, "ranges", ()):
        if (
            merged_range.min_row <= best_row <= merged_range.max_row
            and merged_range.max_row - best_row <= 3
        ):
            header_end = max(header_end, int(merged_range.max_row))
    return best_row, header_end


def _header_candidate_score(row: Sequence[Any]) -> int:
    """让包含多个稳定字段名的行优先于标题行、日期行和普通数据行。"""

    values = [_normalize_header_value(value) for value in row]
    values = [value for value in values if value]
    if not values:
        return 0
    marker_hits = sum(
        1
        for value in values
        if any(marker in value for marker in _HEADER_MARKERS)
    )
    if marker_hits == 0:
        return 0
    numeric_values = sum(1 for value in values if _is_number_value(value))
    return marker_hits * 20 + len(values) - numeric_values * 4


def _first_non_empty_row_index(
    rows: Iterable[Sequence[Any]],
) -> int:
    """返回第一个非空行的 0-based 下标；找不到时返回 0。"""
    for row_index, row in enumerate(rows):
        if _is_nonempty_row(row):
            return row_index

    return 0


def read_headers(
    worksheet: Any,
    header_row: int,
    *,
    header_start_row: int | None = None,
) -> list[str]:
    """读取并去重 Excel 表头，合并父级与子级字段名。"""

    start_row = header_start_row or header_row
    rows = [
        list(row)
        for row in worksheet.iter_rows(
            min_row=start_row,
            max_row=header_row,
            values_only=True,
        )
    ]
    if not rows:
        return []

    last_non_empty = max(
        (
            index
            for row in rows
            for index, value in enumerate(row, start=1)
            if _normalize_header_value(value)
        ),
        default=0,
    )
    for merged_range in getattr(worksheet.merged_cells, "ranges", ()):
        if merged_range.max_row < start_row or merged_range.min_row > header_row:
            continue
        last_non_empty = max(last_non_empty, int(merged_range.max_col))
        value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        for row_number in range(
            max(start_row, merged_range.min_row),
            min(header_row, merged_range.max_row) + 1,
        ):
            matrix_row = rows[row_number - start_row]
            if len(matrix_row) < merged_range.max_col:
                matrix_row.extend([None] * (merged_range.max_col - len(matrix_row)))
            for column_number in range(merged_range.min_col, merged_range.max_col + 1):
                matrix_row[column_number - 1] = value

    # 部分旧版 XLS 转换后，跨列父表头的最后一列可能没有保留在 merged_cells 中。
    # 在同一表头行内，仅对两个已命名字段之间的空白列继承左侧父字段；尾部空白不扩展。
    for row_index, row in enumerate(rows):
        if len(row) < last_non_empty:
            row.extend([None] * (last_non_empty - len(row)))
        last_value: Any | None = None
        for column_index in range(last_non_empty):
            normalized = _normalize_header_value(row[column_index])
            if normalized:
                last_value = row[column_index]
                continue
            has_named_cell_to_the_right = any(
                _normalize_header_value(value)
                for value in row[column_index + 1 : last_non_empty]
            )
            has_child_header = any(
                column_index < len(child_row)
                and bool(_normalize_header_value(child_row[column_index]))
                for child_row in rows[row_index + 1 :]
            )
            if last_value is not None and (
                has_named_cell_to_the_right or has_child_header
            ):
                row[column_index] = last_value

    if last_non_empty == 0:
        return []

    raw_headers: list[str] = []
    for column_number in range(1, last_non_empty + 1):
        parts: list[str] = []
        for row in rows:
            value = row[column_number - 1] if column_number <= len(row) else None
            normalized = _normalize_header_value(value)
            if normalized and normalized not in parts:
                parts.append(normalized)
        raw_headers.append(" / ".join(parts) or f"列{column_number}")
    return _deduplicate_headers(raw_headers)


def _deduplicate_headers(headers: Sequence[str]) -> list[str]:
    used: dict[str, int] = {}
    result: list[str] = []
    for header in headers:
        used[header] = used.get(header, 0) + 1
        result.append(header if used[header] == 1 else f"{header}_{used[header]}")
    return result


def _read_headers_from_values(row: Sequence[Any]) -> list[str]:
    last_non_empty = 0

    for index, value in enumerate(row, start=1):
        if _normalize_header_value(value):
            last_non_empty = index

    if last_non_empty == 0:
        return []

    used: dict[str, int] = {}
    headers: list[str] = []

    for column_index, value in enumerate(row[:last_non_empty], start=1):
        base_name = _normalize_header_value(value) or f"列{column_index}"
        used[base_name] = used.get(base_name, 0) + 1
        name = (
            base_name
            if used[base_name] == 1
            else f"{base_name}_{used[base_name]}"
        )
        headers.append(name)

    return headers


def build_column_profiles(
    *,
    row_iterable: Iterable[Sequence[Any]],
    sheet_id: str,
    headers: list[str],
) -> list[ColumnProfile]:
    """采样前若干个非空值，推断列类型并形成稳定 column_id。"""
    values_by_column: list[list[Any]] = [[] for _ in headers]
    rows_seen = 0

    for row in row_iterable:
        if rows_seen >= MAX_PROFILE_SAMPLE_ROWS:
            break

        if not _is_nonempty_row(row):
            continue

        rows_seen += 1

        for column_index in range(len(headers)):
            value = row[column_index] if column_index < len(row) else None

            if _is_empty(value):
                continue

            values_by_column[column_index].append(value)

    columns: list[ColumnProfile] = []

    for column_index, header in enumerate(headers, start=1):
        values = values_by_column[column_index - 1]
        columns.append(
            ColumnProfile(
                column_id=f"{sheet_id}_col_{column_index}",
                column_index=column_index,
                name=header,
                value_type=infer_column_type(values),
                non_empty_count=len(values),
                sample_values=[
                    _display_value(value)
                    for value in values[:MAX_SAMPLE_VALUES_PER_COLUMN]
                ],
            )
        )

    return columns


def infer_column_type(values: Sequence[Any]) -> ColumnType:
    """按采样值多数推断一列的基础类型。"""
    non_empty = [value for value in values if not _is_empty(value)]

    if not non_empty:
        return ColumnType.UNKNOWN

    if all(isinstance(value, bool) for value in non_empty):
        return ColumnType.BOOLEAN

    if all(isinstance(value, (datetime, date, time)) for value in non_empty):
        return ColumnType.DATE

    if _ratio(non_empty, _is_number_value) >= 0.8:
        return ColumnType.NUMBER

    return ColumnType.STRING


def _read_delimited_rows(*, file_path: Path, suffix: str) -> list[list[str]]:
    """读取 CSV/TSV 行；TSV 固定制表符，CSV 尝试嗅探常见分隔符。"""

    with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)

        if suffix == ".tsv":
            return [list(row) for row in csv.reader(handle, delimiter="\t")]

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel

        return [list(row) for row in csv.reader(handle, dialect)]


def _count_nonempty_rows(rows: Iterable[Sequence[Any]]) -> int:
    return sum(1 for row in rows if _is_nonempty_row(row))


def _is_nonempty_row(row: Sequence[Any]) -> bool:
    return any(not _is_empty(value) for value in row)


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _normalize_header_value(value: Any) -> str:
    return str(value or "").strip()


def _display_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")

    if isinstance(value, (date, time)):
        return value.isoformat()

    return str(value).strip()


def _ratio(values: Sequence[Any], predicate) -> float:
    if not values:
        return 0.0

    return sum(1 for value in values if predicate(value)) / len(values)


def _is_number_value(value: Any) -> bool:
    if isinstance(value, bool):
        return False

    if isinstance(value, (int, float, Decimal)):
        return True

    if isinstance(value, str):
        return _to_decimal(value) is not None

    return False


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None

    text = str(value).strip().replace(",", "").replace("，", "")
    text = text.replace("￥", "").replace("¥", "")

    if not text:
        return None

    try:
        return Decimal(text)
    except InvalidOperation:
        return None
