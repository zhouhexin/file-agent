"""表格工作台测试，保护 Profile、公式校验和只读边界。"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from app.modules.spreadsheet_analysis.profiler import profile_workbook
from app.modules.spreadsheet_workbench.service import SpreadsheetWorkbenchService


def test_profile_workbook_supports_tsv(tmp_path: Path) -> None:
    """TSV 必须进入统一表格 Profile，而不是被当成普通文本。"""

    path = tmp_path / "资助汇总.tsv"
    path.write_text("教师\t资助金额\n张三\t100\n李四\t200\n", encoding="utf-8")

    profile = profile_workbook(
        document_id="doc-tsv",
        filename=path.name,
        file_path=path,
    )

    assert profile.sheets[0].sheet_name == "TSV"
    assert profile.sheets[0].row_count == 2
    assert [column.name for column in profile.sheets[0].columns] == ["教师", "资助金额"]
    assert profile.sheets[0].columns[1].value_type == "number"


def test_workbench_profile_returns_safe_sheet_schema(tmp_path: Path) -> None:
    """profile-spreadsheet 输出只能包含安全结构摘要，不包含本地路径。"""

    path = tmp_path / "科研成果.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "明细"
    worksheet.append(["教师", "资助金额"])
    worksheet.append(["张三", 100])
    workbook.save(path)

    result = SpreadsheetWorkbenchService().profile(
        document_id="doc-xlsx",
        filename=path.name,
        file_path=path,
    )

    assert result["ok"] is True
    assert result["kind"] == "spreadsheet_profile"
    assert result["sheets"][0]["sheet_name"] == "明细"
    assert result["sheets"][0]["columns"][0]["name"] == "教师"
    assert "file_path" not in result


def test_workbench_validate_detects_formula_error_literals(tmp_path: Path) -> None:
    """校验工具必须定位显式公式错误，返回 Sheet 和单元格。"""

    path = tmp_path / "公式错误.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "汇总"
    worksheet.append(["项目", "金额", "公式"])
    worksheet.append(["A", 100, "=SUM(#REF!)"])
    worksheet["D2"] = "#DIV/0!"
    workbook.save(path)

    result = SpreadsheetWorkbenchService().validate(
        document_id="doc-error",
        filename=path.name,
        file_path=path,
    )

    errors = {(item["sheet_name"], item["cell"], item["error"]) for item in result["formula_errors"]}
    assert result["ok"] is True
    assert result["status"] == "NEEDS_REVIEW"
    assert ("汇总", "C2", "#REF!") in errors
    assert ("汇总", "D2", "#DIV/0!") in errors


def test_workbench_validate_marks_xlsm_without_running_macro(tmp_path: Path) -> None:
    """xlsm 文件必须标记宏风险，但校验过程不能执行宏。"""

    path = tmp_path / "含宏表.xlsm"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["字段"])
    worksheet.append(["值"])
    workbook.save(path)

    result = SpreadsheetWorkbenchService().validate(
        document_id="doc-xlsm",
        filename=path.name,
        file_path=path,
    )

    assert result["status"] == "NEEDS_REVIEW"
    assert result["warnings"][0]["code"] == "MACRO_FILE_DETECTED"
