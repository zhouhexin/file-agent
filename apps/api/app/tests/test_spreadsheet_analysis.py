"""表格分析链路测试，保护受控查询计划、确定性执行和 Planner 路由边界。"""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

from app.modules.agent.planner import DeterministicPlanner
from app.modules.spreadsheet_analysis.executor import execute_query
from app.modules.spreadsheet_analysis.formatter import format_spreadsheet_analysis_response
from app.modules.spreadsheet_analysis.profiler import profile_workbook
from app.modules.spreadsheet_analysis.query_planner import build_deterministic_query_plans
from app.modules.spreadsheet_analysis.schemas import SpreadsheetQueryPlan
from app.modules.spreadsheet_analysis.service import SpreadsheetAnalysisService
from app.modules.spreadsheet_analysis.validator import SpreadsheetPlanValidationError, validate_plan


class FakeJsonClient:
    """测试用 LLM JSON 客户端，固定返回受控查询计划。"""

    def __init__(self, response: dict) -> None:
        """保存测试指定的模型响应。"""

        self.response = response

    def complete_json(self, *, system_prompt: str, user_payload: dict) -> dict:
        """返回固定 JSON，避免测试依赖真实外部模型。"""

        return self.response


def _make_workbook(tmp_path: Path) -> Path:
    """创建包含人员、论文类型和资助金额的临时工作簿。"""

    path = tmp_path / "科研成果汇总.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "汇总表"
    worksheet.append(["申请人", "论文类型", "资助类别", "资助金额"])
    worksheet.append(["张三", "CCF A类论文", "重点", 1000])
    worksheet.append(["李四", "CCF A类论文", "重点", 2000])
    worksheet.append(["王五", "核心期刊论文", "一般", 500])
    workbook.save(path)
    return path


def _profile(path: Path):
    """读取临时工作簿 Profile，用于后续计划校验和执行。"""

    return profile_workbook(
        document_id="doc-1",
        filename=path.name,
        file_path=path,
    )


def test_grouped_sum_executes_without_business_keyword_rules(tmp_path: Path) -> None:
    """分组求和必须由受控计划和执行器完成，不能依赖业务关键词硬编码。"""

    path = _make_workbook(tmp_path)
    profile = _profile(path)
    plan = SpreadsheetQueryPlan.model_validate(
        {
            "sheet_id": "sheet_1",
            "metric": {
                "operation": "sum",
                "column_id": "sheet_1_col_4",
                "label": "资助金额合计",
            },
            "group_by_column_id": "sheet_1_col_2",
            "filters": [],
            "sort_direction": "desc",
            "limit": 50,
        }
    )

    result = execute_query(
        file_path=path,
        profile=profile,
        plan=validate_plan(profile=profile, plan=plan),
    )

    assert result["status"] == "COMPLETED"
    assert result["rows_scanned"] == 3
    assert result["rows_included"] == 3
    assert result["results"] == [
        {"group": "CCF A类论文", "value": "3000"},
        {"group": "核心期刊论文", "value": "500"},
    ]


def test_count_rows_with_filter(tmp_path: Path) -> None:
    """带筛选条件的计数必须只扫描受控列和受控操作。"""

    path = _make_workbook(tmp_path)
    profile = _profile(path)
    plan = SpreadsheetQueryPlan.model_validate(
        {
            "sheet_id": "sheet_1",
            "metric": {"operation": "count_rows", "label": "论文数量"},
            "filters": [
                {
                    "column_id": "sheet_1_col_2",
                    "operator": "equals",
                    "value": "CCF A类论文",
                }
            ],
        }
    )

    result = execute_query(
        file_path=path,
        profile=profile,
        plan=validate_plan(profile=profile, plan=plan),
    )

    assert result["results"] == [{"group": "全部", "value": "2"}]
    assert result["rows_matched"] == 2


def test_tsv_sum_executes_with_same_query_pipeline(tmp_path: Path) -> None:
    """TSV 必须复用统一表格分析链路，而不是退回普通文本读取。"""

    path = tmp_path / "资助汇总.tsv"
    path.write_text("教师\t资助金额\n张三\t100\n李四\t200\n", encoding="utf-8")
    profile = _profile(path)
    plan = SpreadsheetQueryPlan.model_validate(
        {
            "sheet_id": "sheet_1",
            "metric": {
                "operation": "sum",
                "column_id": "sheet_1_col_2",
                "label": "资助金额合计",
            },
        }
    )

    result = execute_query(
        file_path=path,
        profile=profile,
        plan=validate_plan(profile=profile, plan=plan),
    )

    assert result["status"] == "COMPLETED"
    assert result["results"] == [{"group": "全部", "value": "300"}]


def test_validator_rejects_hallucinated_column_id(tmp_path: Path) -> None:
    """校验器必须拒绝不存在的 column_id，防止 LLM 编造字段。"""

    path = _make_workbook(tmp_path)
    profile = _profile(path)
    plan = SpreadsheetQueryPlan.model_validate(
        {
            "sheet_id": "sheet_1",
            "metric": {"operation": "sum", "column_id": "sheet_1_col_999"},
        }
    )

    with pytest.raises(SpreadsheetPlanValidationError):
        validate_plan(profile=profile, plan=plan)


def test_service_returns_clarification_without_executing(tmp_path: Path) -> None:
    """LLM 要求澄清时服务只能返回可选字段，不能继续执行表格查询。"""

    path = _make_workbook(tmp_path)
    service = SpreadsheetAnalysisService(
        settings=SimpleNamespace(llm_enabled=True),
        client=FakeJsonClient(
            {
                "clarification_required": True,
                "clarification_question": "你希望按哪一列汇总？",
            }
        ),
    )

    result = service.analyze(
        document_id="doc-1",
        filename=path.name,
        file_path=path,
        question="汇总成果",
    )

    assert result["ok"] is True
    assert result["status"] == "NEEDS_CLARIFICATION"
    assert "论文类型" in result["available_sheets"][0]["columns"]


def test_remediation_status_statistics_include_requested_zero_categories(
    tmp_path: Path,
) -> None:
    """整改状态统计必须补齐零值类别，并把“正在整改”归入用户所说的“持续整改”。"""

    path = tmp_path / "审计整改自查表.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "整改事项"
    worksheet.append(
        ["序号", "问题描述", "整改完成情况（已整改完成、正在整改、未整改）"]
    )
    worksheet.append([1, "问题一", "已整改完成"])
    worksheet.append([2, "问题二", "已整改完成"])
    worksheet.append([3, "问题三", "已整改完成"])
    workbook.save(path)

    result = SpreadsheetAnalysisService(
        settings=SimpleNamespace(llm_enabled=True),
        client=FakeJsonClient(
            {
                "sheet_id": "sheet_1",
                "metric": {"operation": "count_rows", "label": "事项数量"},
                "group_by_column_id": "sheet_1_col_3",
            }
        ),
    ).analyze(
        document_id="doc-audit-remediation",
        filename=path.name,
        file_path=path,
        question="从审计整改自查表中统计已整改、未整改和持续整改事项",
    )
    response = format_spreadsheet_analysis_response([result])

    assert result["results"] == [
        {"group": "已整改", "value": "3"},
        {"group": "未整改", "value": "0"},
        {"group": "持续整改", "value": "0"},
    ]
    assert "“持续整改”按表内“正在整改/整改中”口径统计。" in result["warnings"]
    assert "- 已整改：3" in response
    assert "- 未整改：0" in response
    assert "- 持续整改：0" in response


def test_deterministic_planner_routes_uploaded_xlsx_to_spreadsheet_tool() -> None:
    """上传 Excel 后的统计请求必须路由到只读表格分析 Tool。"""

    plan = DeterministicPlanner().plan(
        conversation_id="conversation-1",
        user_id="user-1",
        message_id="message-1",
        message="按论文类型统计成果数量",
        attachments=[{"document_id": "doc-1", "filename": "科研成果汇总.xlsx"}],
    )

    assert plan.intent == "ANALYZE_SPREADSHEET"
    assert plan.steps[0].tool_name == "analyze-spreadsheet"


def test_person_total_amount_uses_deterministic_multi_sheet_plan_without_llm(tmp_path: Path) -> None:
    """“某人的总金额”必须按真实列筛选并跨 Sheet 求和，不调用 LLM 计算或猜数。"""

    path = tmp_path / "2024科研成果资助汇总表.xlsx"
    workbook = openpyxl.Workbook()
    paper = workbook.active
    paper.title = "论文"
    paper.append(["序号", "申请人", "资助金额"])
    paper.append([1, "张三", 100])
    paper.append([2, "金海燕", 3000])
    paper.append([3, "金海燕", 2000])
    patent = workbook.create_sheet("专利")
    patent.append(["序号", "申请人", "资助金额"])
    patent.append([1, "金海燕", 1500])
    workbook.save(path)

    profile = profile_workbook(
        document_id="doc-person-total",
        filename=path.name,
        file_path=path,
    )
    plans = build_deterministic_query_plans(
        question="金海燕的资助总金额是多少",
        profile=profile,
    )
    scoped_plans = build_deterministic_query_plans(
        question="2024科研成果资助汇总表中金海燕的资助总金额是多少",
        profile=profile,
    )
    fallback_plans = build_deterministic_query_plans(
        question="2024科研成果资助汇总表中欧阳娜娜的资助总金额是多少",
        profile=profile,
    )

    assert [plan.sheet_id for plan in plans] == ["sheet_1", "sheet_2"]
    assert all(plan.metric and plan.metric.operation.value == "sum" for plan in plans)
    assert all(plan.filters[0].value == "金海燕" for plan in plans)
    assert [plan.sheet_id for plan in scoped_plans] == ["sheet_1", "sheet_2"]
    assert all(plan.filters[0].value == "金海燕" for plan in scoped_plans)
    # Profile 样本未包含目标人员时，也必须先剥离文件范围，不能恢复旧的长标题误识别。
    assert [plan.sheet_id for plan in fallback_plans] == ["sheet_1", "sheet_2"]
    assert all(plan.filters[0].value == "欧阳娜娜" for plan in fallback_plans)

    result = SpreadsheetAnalysisService(
        settings=SimpleNamespace(llm_enabled=False),
    ).analyze(
        document_id="doc-person-total",
        filename=path.name,
        file_path=path,
        question="金海燕的资助总金额是多少",
    )
    response = format_spreadsheet_analysis_response([result])

    assert result["results"] == [{"group": "全部", "value": "6500"}]
    assert result["rows_matched"] == 3
    assert [item["sheet_name"] for item in result["sheet_breakdown"]] == ["论文", "专利"]
    assert "资助金额合计为 6,500" in response
    assert "Sheet“论文” / “资助金额”：5,000" in response
    assert "Sheet“专利” / “资助金额”：1,500" in response
    assert "计算方式：5,000 + 1,500 = 6,500" in response
    assert " B3" not in response

    scoped_result = SpreadsheetAnalysisService(
        settings=SimpleNamespace(llm_enabled=False),
    ).analyze(
        document_id="doc-person-total",
        filename=path.name,
        file_path=path,
        question="2024科研成果资助汇总表中金海燕的资助总金额是多少",
    )
    assert scoped_result["results"] == [{"group": "全部", "value": "6500"}]
    assert scoped_result["filters"][0]["value"] == "金海燕"


def test_job_quantity_uses_merged_multi_level_headers_and_sums_job_columns(
    tmp_path: Path,
) -> None:
    """“岗位数量”必须汇总招聘计划数值列，不能把学科行数当成岗位数。"""

    path = tmp_path / "2026年师资招聘计划申报表.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "专任教师岗位"
    worksheet["A1"] = "附件4："
    worksheet.merge_cells("A2:H2")
    worksheet["A2"] = "西安理工大学2026年师资招聘计划申报表"
    worksheet.append(["单位（盖章）：", None, None, None, None, None, None, "填报日期：2026年1月5日"])
    worksheet.append(["序号", "拟补充学科", "拟补充系所", "现有人数", "招聘计划", None, None, None])
    worksheet.append([None, None, None, None, "拟补充方向（团队）", "优秀人才", "青年人才（事业编制）", "青年人才（预聘制）"])
    for merged_range in ("A4:A5", "B4:B5", "C4:C5", "D4:D5", "E4:G4"):
        worksheet.merge_cells(merged_range)
    rows = [
        [1, "计算机科学与技术", "计算机科学与技术系", 27, "智能信息处理", 2, None, 1],
        [2, None, "软件工程系", 25, "智能图像处理", 1, None, 1],
        [3, None, "物联网工程系", 27, "物联网系统", 1, None, 1],
        [4, None, "人工智能系", 15, "人工智能", 2, None, 1],
        [5, "网络空间安全", "网络空间安全系", 27, "信息安全", 1, None, 1],
    ]
    for row_number, row in enumerate(rows, start=6):
        worksheet.append(row)
        worksheet.merge_cells(start_row=row_number, start_column=6, end_row=row_number, end_column=7)
    worksheet["A13"] = "填报说明：招聘计划人数按类别填报。"
    worksheet.merge_cells("A13:H13")
    workbook.save(path)

    profile = profile_workbook(
        document_id="doc-recruitment",
        filename=path.name,
        file_path=path,
    )
    assert profile.sheets[0].header_row == 5
    assert profile.sheets[0].columns[5].name == "招聘计划 / 优秀人才"
    assert profile.sheets[0].columns[7].name == "招聘计划 / 青年人才（预聘制）"

    plans = build_deterministic_query_plans(
        question="统计2026年师资招聘计划申报表中的岗位数量",
        profile=profile,
    )
    assert [plan.metric.column_id for plan in plans if plan.metric] == [
        "sheet_1_col_6",
        "sheet_1_col_8",
    ]
    assert all(plan.metric and plan.metric.operation.value == "sum" for plan in plans)

    result = SpreadsheetAnalysisService(
        settings=SimpleNamespace(llm_enabled=False),
    ).analyze(
        document_id="doc-recruitment",
        filename=path.name,
        file_path=path,
        question="统计2026年师资招聘计划申报表中的岗位数量",
    )
    response = format_spreadsheet_analysis_response([result])

    assert result["results"] == [{"group": "全部", "value": "12"}]
    assert [item["value"] for item in result["sheet_breakdown"]] == ["7", "5"]
    assert result["warnings"] == []
    assert "岗位总数为 12 个" in response
    assert "7 + 5 = 12" in response


def test_job_quantity_rejects_llm_row_count_substitution(tmp_path: Path) -> None:
    """没有可确认的岗位数值列时，LLM 不得用 count_rows 冒充岗位数量。"""

    path = _make_workbook(tmp_path)
    result = SpreadsheetAnalysisService(
        settings=SimpleNamespace(llm_enabled=True),
        client=FakeJsonClient(
            {
                "sheet_id": "sheet_1",
                "metric": {"operation": "count_rows", "label": "岗位数量"},
            }
        ),
    ).analyze(
        document_id="doc-1",
        filename=path.name,
        file_path=path,
        question="统计岗位数量",
    )

    assert result["status"] == "NEEDS_CLARIFICATION"
    assert "不能按数据行数代替" in result["message"]
