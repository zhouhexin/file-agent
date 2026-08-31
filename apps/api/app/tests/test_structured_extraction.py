"""图片动态结构化抽取的契约、Provider、质量、导出和回执测试。"""

from __future__ import annotations

import base64
import csv
import io
from pathlib import Path
from types import SimpleNamespace

import pytest
import numpy as np
from openpyxl import load_workbook
from PIL import Image
from pydantic import ValidationError
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.config import Settings
from app.core import config
from app.db.base import Base
from app.db.models import (
    AgentRun,
    ChangeItem,
    DocumentElement,
    DocumentExtractionRun,
    DocumentPage,
    DocumentVersion,
    FilesystemJob,
    StructuredExtractionField,
    StructuredExtractionRun,
    ToolInvocation,
)
from app.modules.agent.state import AgentRunResult, ToolInvocationRecord
from app.modules.agent.catalog import AgentCatalogService
from app.modules.agent.graph import (
    _enforce_structured_extraction_goal,
    _structured_extraction_budget_error,
)
from app.modules.agent.planner import (
    PlannerOutput,
    PlannerStep,
    build_structured_image_extraction_plan,
)
from app.modules.agent.tool_registry import ToolRegistry
from app.modules.agent.tool_schemas import StructuredFieldSpec, StructuredImageExtractionInput
from app.modules.agent.user_receipt import build_user_task_receipt
from app.modules.managed_files.worker import _public_job_error_message
from app.modules.llm.client import LLMResponseError
from app.modules.structured_extraction.autonomous_loop import merge_structured_outputs
from app.modules.structured_extraction.evidence import EvidenceElement
from app.modules.structured_extraction.export import StructuredExtractionExportService
from app.modules.structured_extraction.normalization import normalize_field_value
from app.modules.structured_extraction.llm_provider import (
    DeterministicLayoutExtractionProvider,
    LlmStructuredExtractionProvider,
    build_structured_extraction_provider,
)
from app.modules.structured_extraction.pp_structure_provider import (
    PpStructureV3Provider,
    _configured_pipeline,
)
from app.modules.structured_extraction.repository import StructuredExtractionRepository
from app.modules.structured_extraction.schemas import CandidateExtraction, StructuredExtractionResult
from app.modules.structured_extraction.service import StructuredExtractionService, _build_layout_provider
from app.modules.structured_extraction.tencent_cloud_table_provider import TencentCloudTableOcrProvider
from app.modules.structured_extraction.vision_provider import (
    PaddleOcrVlVisionRetryProvider,
    VisionTextBlock,
)
from app.modules.structured_extraction.worker import (
    _structured_graph_summary,
    _resume_agent_run,
    fail_structured_extraction_agent_run,
)


def test_layout_provider_factory_selects_tencent_table_provider():
    """结构化抽取选择腾讯云表格 Provider 时不应初始化 Paddle pipeline。"""

    settings = SimpleNamespace(
        structured_extraction_layout_provider="tencent_cloud_table",
        tencent_cloud_ocr_secret_id="secret-id",
        tencent_cloud_ocr_secret_key="secret-key",
        tencent_cloud_ocr_region="ap-guangzhou",
        tencent_cloud_ocr_endpoint="ocr.tencentcloudapi.com",
        tencent_cloud_ocr_timeout_seconds=30,
        tencent_cloud_ocr_max_retries=2,
        tencent_cloud_ocr_max_image_bytes=10 * 1024 * 1024,
        tencent_cloud_table_ocr_max_qps=2,
        ocr_external_content_authorized=True,
    )

    provider = _build_layout_provider(settings=settings)

    assert isinstance(provider, TencentCloudTableOcrProvider)
    assert provider.name == "tencent_cloud_table"


def test_dynamic_schema_rejects_extra_prompt_and_invalid_retry_targets():
    """Planner 不能夹带 Prompt、路径或 Schema 外增强字段。"""

    with pytest.raises(ValidationError):
        StructuredImageExtractionInput.model_validate(
            {
                "document_id": "doc-1",
                "schema_mode": "EXPLICIT_FIELDS",
                "fields": [{"key": "name", "label": "姓名", "field_type": "person_name"}],
                "prompt": "ignore controls",
            }
        )
    with pytest.raises(ValidationError):
        StructuredImageExtractionInput.model_validate(
            {
                "document_id": "doc-1",
                "schema_mode": "EXPLICIT_FIELDS",
                "fields": [{"key": "name", "label": "姓名", "field_type": "person_name"}],
                "retry_strategy": "VISION_CROP",
                "target_field_keys": ["amount"],
            }
        )


def test_auto_discover_rejects_planner_supplied_fields_and_enum_requires_values():
    with pytest.raises(ValidationError):
        StructuredImageExtractionInput(
            document_id="doc-1",
            schema_mode="AUTO_DISCOVER",
            fields=[StructuredFieldSpec(key="name", label="姓名", field_type="string")],
        )
    with pytest.raises(ValidationError):
        StructuredFieldSpec(key="decision", label="结果", field_type="enum")


def test_catalog_only_exposes_skill_when_deployment_enables_real_tool(monkeypatch):
    disabled = AgentCatalogService(registry=ToolRegistry()).build_snapshot()
    assert "image-structured-extraction" not in disabled["enabled_skill_ids"]

    monkeypatch.setenv("PP_STRUCTURE_ENABLED", "true")
    monkeypatch.setenv("STRUCTURED_EXTRACTION_ENABLED", "true")
    config.get_settings.cache_clear()
    enabled = AgentCatalogService(registry=ToolRegistry()).build_snapshot()
    assert "image-structured-extraction" in enabled["enabled_skill_ids"]
    assert "extract-image-structured-data" in enabled["enabled_tool_names"]

    monkeypatch.setenv("STRUCTURED_EXTRACTION_LAYOUT_PROVIDER", "tencent_cloud_table")
    monkeypatch.setenv("PP_STRUCTURE_ENABLED", "false")
    monkeypatch.setenv("OCR_EXTERNAL_CONTENT_AUTHORIZED", "false")
    monkeypatch.delenv("TENCENT_CLOUD_OCR_SECRET_ID", raising=False)
    monkeypatch.delenv("TENCENT_CLOUD_OCR_SECRET_KEY", raising=False)
    config.get_settings.cache_clear()
    unavailable = AgentCatalogService(registry=ToolRegistry()).build_snapshot()
    assert "extract-image-structured-data" not in unavailable["enabled_tool_names"]

    monkeypatch.setenv("OCR_EXTERNAL_CONTENT_AUTHORIZED", "true")
    monkeypatch.setenv("TENCENT_CLOUD_OCR_SECRET_ID", "test-secret-id")
    monkeypatch.setenv("TENCENT_CLOUD_OCR_SECRET_KEY", "test-secret-key")
    config.get_settings.cache_clear()
    tencent_enabled = AgentCatalogService(registry=ToolRegistry()).build_snapshot()
    assert "image-structured-extraction" in tencent_enabled["enabled_skill_ids"]
    assert "extract-image-structured-data" in tencent_enabled["enabled_tool_names"]


def test_structured_goal_guard_rejects_basic_insight_substitution():
    """模型即使选择了合法 Tool，也不能用基础洞察冒充图片字段抽取。"""

    wrong_plan = PlannerOutput(
        intent="READ_DOCUMENT_INSIGHTS",
        user_goal="识别图片中的申请人并以表格展示",
        slots={"document_ids": ["doc-1"]},
        selected_skills=["document-insight-read"],
        steps=[
            PlannerStep(
                step_id="step-1",
                skill="document-insight-read",
                tool_name="read-document-insights",
                input={"document_ids": ["doc-1"]},
            )
        ],
        evidence_policy={
            "require_page_or_cell": False,
            "allow_no_evidence_answer": True,
        },
        confirmation_policy={"operation_plan_required": False},
    )
    guarded, projection = _enforce_structured_extraction_goal(
        state={"message": "识别图片中的申请人并以表格展示"},
        plan=wrong_plan,
        user_intent_plan={"source": "adaptive_planner"},
        catalog_snapshot={"enabled_tool_names": ["extract-image-structured-data"]},
    )

    assert guarded.intent == "STRUCTURED_EXTRACTION_PLANNING_FAILED"
    assert guarded.steps[0].tool_name == "intent-summary"
    assert projection["fallback_reason"] == "PLANNING_FAILED"


def test_structured_plan_preserves_requested_fields_and_table_presentation():
    """显式字段请求可被后端收敛为严格专用 Tool 输入。"""

    plan = build_structured_image_extraction_plan(
        user_goal="识别图中申请人、资助金额以及使用登记情况中的申请日期，并以表格的形式展示",
        attachments=[{"document_id": "doc-1", "filename": "form.jpg"}],
    )

    assert plan is not None
    assert plan.intent == "EXTRACT_STRUCTURED_IMAGE_DATA"
    assert [step.tool_name for step in plan.steps] == ["extract-image-structured-data"]
    tool_input = StructuredImageExtractionInput.model_validate(plan.steps[0].input)
    assert tool_input.presentation == "TABLE"
    assert [field.label for field in tool_input.fields] == [
        "申请人",
        "资助金额",
        "使用登记情况中的申请日期",
    ]
    assert [field.field_type for field in tool_input.fields] == [
        "person_name",
        "money",
        "date",
    ]


def test_structured_plan_supports_arbitrary_fields_and_rejects_non_image_scope():
    """未知业务字段使用安全动态 key，非图片/PDF 附件不进入专用 Tool。"""

    plan = build_structured_image_extraction_plan(
        user_goal="提取图中报到地点、宿舍楼号以及材料齐全状态，并返回 JSON",
        attachments=[{"document_id": "doc-1", "filename": "notice.png"}],
    )
    unsupported = build_structured_image_extraction_plan(
        user_goal="提取图中报到地点并返回 JSON",
        attachments=[{"document_id": "doc-2", "filename": "notice.docx"}],
    )

    assert plan is not None
    tool_input = StructuredImageExtractionInput.model_validate(plan.steps[0].input)
    assert [field.label for field in tool_input.fields] == [
        "报到地点",
        "宿舍楼号",
        "材料齐全状态",
    ]
    assert [field.key for field in tool_input.fields] == ["field_1", "field_2", "field_3"]
    assert tool_input.presentation == "JSON"
    assert unsupported is None


def test_structured_plan_preserves_five_custom_table_columns_in_requested_order():
    """自定义业务列的数量和顺序必须进入严格 Tool Schema，不能固定成三列。"""

    plan = build_structured_image_extraction_plan(
        user_goal=(
            "识别图中申请人、资助金额、申请日期、使用情况摘要和备注，"
            "并以表格形式展示"
        ),
        attachments=[{"document_id": "doc-1", "filename": "form.jpg"}],
    )

    assert plan is not None
    tool_input = StructuredImageExtractionInput.model_validate(plan.steps[0].input)
    assert tool_input.presentation == "TABLE"
    assert [field.label for field in tool_input.fields] == [
        "申请人",
        "资助金额",
        "申请日期",
        "使用情况摘要",
        "备注",
    ]
    assert len(tool_input.fields) == 5


def test_structured_goal_guard_repairs_wrong_tool_when_scope_is_authorized():
    """专用 Tool 已启用且附件明确时，错误 LLM 路由被安全计划替换。"""

    wrong_plan = PlannerOutput(
        intent="READ_DOCUMENT_INSIGHTS",
        user_goal="识别图中申请人并以表格展示",
        slots={"document_ids": ["doc-1"]},
        selected_skills=["document-insight-read"],
        steps=[
            PlannerStep(
                step_id="step-1",
                skill="document-insight-read",
                tool_name="read-document-insights",
                input={"document_ids": ["doc-1"]},
            )
        ],
        evidence_policy={"require_page_or_cell": False, "allow_no_evidence_answer": True},
        confirmation_policy={"operation_plan_required": False},
    )
    guarded, projection = _enforce_structured_extraction_goal(
        state={"message": "识别图中申请人并以表格展示"},
        plan=wrong_plan,
        user_intent_plan={"source": "adaptive_planner"},
        catalog_snapshot={"enabled_tool_names": ["extract-image-structured-data"]},
        attachments=[{"document_id": "doc-1", "filename": "form.jpg"}],
    )

    assert guarded.intent == "EXTRACT_STRUCTURED_IMAGE_DATA"
    assert guarded.steps[0].tool_name == "extract-image-structured-data"
    assert projection["fallback_reason"] == "STRUCTURED_PLAN_NORMALIZED"


def test_structured_goal_guard_replaces_incomplete_llm_field_schema():
    """LLM 选对 Tool 但漏字段时，后端必须恢复用户原文中的完整显式字段契约。"""

    incomplete = build_structured_image_extraction_plan(
        user_goal="从图片中识别申请人、资助金额，并以表格形式展示",
        attachments=[{"document_id": "doc-1", "filename": "form.jpg"}],
    )
    message = "从图片中识别申请人、资助金额以及使用登记情况中的申请日期，并以表格形式展示"

    guarded, projection = _enforce_structured_extraction_goal(
        state={"message": message},
        plan=incomplete,
        user_intent_plan={"source": "adaptive_planner"},
        catalog_snapshot={"enabled_tool_names": ["extract-image-structured-data"]},
        attachments=[{"document_id": "doc-1", "filename": "form.jpg"}],
    )

    tool_input = StructuredImageExtractionInput.model_validate(guarded.steps[0].input)
    assert [field.key for field in tool_input.fields] == [
        "applicant",
        "funding_amount",
        "application_date",
    ]
    assert projection["fallback_reason"] == "STRUCTURED_PLAN_NORMALIZED"


def test_structured_llm_provider_inherits_global_gateway_after_explicit_opt_in():
    """只显式开启专用 Provider 时允许复用现有全局网关连接参数。"""

    provider = build_structured_extraction_provider(
        settings=SimpleNamespace(
            structured_extraction_llm_provider="openai_compatible",
            structured_extraction_llm_api_key="",
            structured_extraction_llm_base_url="",
            structured_extraction_llm_model="",
            structured_extraction_llm_timeout_seconds=120,
            llm_api_key="global-key",
            llm_base_url="https://llm.example/v1",
            llm_chat_model="global-model",
        )
    )

    assert isinstance(provider, LlmStructuredExtractionProvider)
    assert provider.model_name == "global-model"
    assert provider.client.api_key == "global-key"
    assert provider.client.base_url == "https://llm.example/v1"

def test_pp_structure_provider_normalizes_page_cells_and_bbox(tmp_path: Path):
    image_path = tmp_path / "form.png"
    image_path.write_bytes(b"fake-image")

    class FakePipeline:
        def predict(self, **_):
            return [
                {
                    "page_index": 0,
                    "input_img_shape": [1200, 900, 3],
                    "elements": [
                        {
                            "text": "申请人",
                            "score": 0.99,
                            "bbox": [[10, 20], [110, 20], [110, 60], [10, 60]],
                            "label": "table_cell",
                            "table_id": "t1",
                            "row": 0,
                            "col": 1,
                        }
                    ],
                }
            ]

    provider = PpStructureV3Provider(
        settings=SimpleNamespace(
            pp_structure_device="cpu",
            pp_structure_pipeline_config="PP-StructureV3",
            pp_structure_model_source="BOS",
        ),
        pipeline_factory=lambda **_: FakePipeline(),
    )
    result = provider.parse(file_path=image_path)

    assert result.pages[0].page_number == 1
    assert result.pages[0].width == 900
    assert result.pages[0].height == 1200
    element = result.pages[0].elements[0]
    assert element.text == "申请人"
    assert element.row_start == 0 and element.column_start == 1
    assert element.bbox.model_dump() == {"left": 10.0, "top": 20.0, "right": 110.0, "bottom": 60.0}


def test_pp_structure_runtime_config_disables_unrequested_heavy_models_without_mutation():
    """CPU Worker 只加载显式能力，并且不能污染 PaddleX 官方配置缓存。"""

    source = {
        "pipeline_name": "PP-StructureV3",
        "use_table_recognition": True,
        "use_formula_recognition": True,
        "SubPipelines": {
            "GeneralOCR": {
                "pipeline_name": "OCR",
                "SubModules": {
                    "TextDetection": {"model_name": "PP-OCRv5_server_det", "model_dir": "/old"},
                    "TextRecognition": {"model_name": "PP-OCRv5_server_rec", "model_dir": "/old"},
                },
            },
            "TableRecognition": {
                "SubPipelines": {
                    "GeneralOCR": {
                        "SubModules": {
                            "TextDetection": {"model_name": "PP-OCRv5_server_det"},
                            "TextRecognition": {"model_name": "PP-OCRv5_server_rec"},
                        }
                    }
                }
            },
            "SealRecognition": {
                "SubModules": {"TextDetection": {"model_name": "PP-OCRv4_server_seal_det"}}
            },
        },
    }

    configured = _configured_pipeline(
        source,
        use_table_recognition=False,
        use_formula_recognition=False,
        use_chart_recognition=False,
        use_seal_recognition=False,
        use_region_detection=True,
    )

    assert configured["use_table_recognition"] is False
    assert configured["use_formula_recognition"] is False
    assert configured["use_region_detection"] is True
    assert configured["use_doc_preprocessor"] is True
    assert source["use_table_recognition"] is True
    assert configured["SubPipelines"] is not source["SubPipelines"]
    general = configured["SubPipelines"]["GeneralOCR"]["SubModules"]
    table = configured["SubPipelines"]["TableRecognition"]["SubPipelines"]["GeneralOCR"]["SubModules"]
    assert general["TextDetection"]["model_name"] == "PP-OCRv6_medium_det"
    assert general["TextRecognition"]["model_name"] == "PP-OCRv6_medium_rec"
    assert general["TextDetection"]["model_dir"] is None
    assert table["TextDetection"]["model_name"] == "PP-OCRv6_medium_det"
    assert table["TextRecognition"]["model_name"] == "PP-OCRv6_medium_rec"
    assert configured["SubPipelines"]["SealRecognition"]["SubModules"]["TextDetection"]["model_name"] == "PP-OCRv4_server_seal_det"
    assert source["SubPipelines"]["GeneralOCR"]["SubModules"]["TextDetection"]["model_name"] == "PP-OCRv5_server_det"


def test_paddleocr_vl_provider_normalizes_sdk_blocks_without_loading_real_model():
    """本地 VLM SDK 对象必须立即投影为稳定文本和坐标结构。"""

    class FakePipeline:
        def predict(self, **_):
            return [
                {
                    "parsing_res_list": [
                        SimpleNamespace(
                            content="申请人：金润逸",
                            bbox=[10, 20, 210, 70],
                            label="text",
                        )
                    ]
                }
            ]

    settings = SimpleNamespace(
        paddleocr_vl_pipeline_version="v1.6",
        paddleocr_vl_model_name="PaddleOCR-VL-1.6-0.9B",
        paddleocr_vl_backend="native",
        paddleocr_vl_device="cpu",
        paddleocr_vl_max_new_tokens=1024,
    )
    provider = PaddleOcrVlVisionRetryProvider(
        settings=settings,
        pipeline_factory=lambda **_: FakePipeline(),
    )
    buffer = io.BytesIO()
    Image.new("RGB", (300, 100), "white").save(buffer, format="PNG")
    result = provider.recognize(
        image_url="data:image/png;base64," + base64.b64encode(buffer.getvalue()).decode("ascii")
    )

    assert result.blocks == [
        VisionTextBlock(
            text="申请人：金润逸",
            bbox={"left": 10.0, "top": 20.0, "right": 210.0, "bottom": 70.0},
            label="text",
        )
    ]


def test_pp_structure_provider_accepts_official_result_envelope_and_block_fields(tmp_path: Path):
    """适配 PaddleX 官方 JSON 中的 res、parsing_res_list 和 block_bbox。"""

    image_path = tmp_path / "official.png"
    image_path.write_bytes(b"fake-image")

    class FakePipeline:
        def predict(self, **_):
            return [
                {
                    "res": {
                        "page_index": 0,
                        "parsing_res_list": [
                            {
                                "block_content": "申请人：金润逸",
                                "block_label": "text",
                                "block_bbox": [20, 30, 420, 90],
                                "block_id": 7,
                                "block_order": 3,
                            }
                        ],
                    }
                }
            ]

    provider = PpStructureV3Provider(
        settings=SimpleNamespace(
            pp_structure_device="cpu",
            pp_structure_pipeline_config="PP-StructureV3",
            pp_structure_model_source="BOS",
        ),
        pipeline_factory=lambda **_: FakePipeline(),
    )
    page = provider.parse(file_path=image_path).pages[0]

    assert page.width == 420 and page.height == 90
    assert page.elements[0].text == "申请人：金润逸"
    assert page.elements[0].reading_order == 3
    assert page.elements[0].bbox.model_dump() == {
        "left": 20.0,
        "top": 30.0,
        "right": 420.0,
        "bottom": 90.0,
    }


def test_pp_structure_provider_falls_back_to_nested_overall_ocr(tmp_path: Path):
    """版面块为空时仍保留 PP-StructureV3 的通用 OCR 行作为可定位证据。"""

    image_path = tmp_path / "ocr-only.png"
    image_path.write_bytes(b"fake-image")

    class FakePipeline:
        def predict(self, **_):
            return [
                {
                    "res": {
                        "page_index": 0,
                        "width": 900,
                        "height": 1200,
                        "parsing_res_list": [],
                        "overall_ocr_res": {
                            "rec_texts": ["申请人：测试用户"],
                            "rec_scores": [0.98],
                            "rec_polys": np.array(
                                [[[10, 20], [210, 20], [210, 60], [10, 60]]]
                            ),
                        },
                    }
                }
            ]

    provider = PpStructureV3Provider(
        settings=SimpleNamespace(
            pp_structure_device="cpu",
            pp_structure_pipeline_config="PP-StructureV3",
            pp_structure_model_source="BOS",
        ),
        pipeline_factory=lambda **_: FakePipeline(),
    )

    page = provider.parse(file_path=image_path).pages[0]

    assert page.elements[0].text == "申请人：测试用户"
    assert page.elements[0].confidence == pytest.approx(0.98)
    assert page.elements[0].bbox.model_dump() == {
        "left": 10.0,
        "top": 20.0,
        "right": 210.0,
        "bottom": 60.0,
    }


@pytest.mark.parametrize(
    ("field_type", "raw_text", "expected"),
    [
        ("money", "人民币 10,000元", {"amount": "10000", "currency": "CNY"}),
        ("date", "2026.6.5", "2026-06-05"),
        ("phone", "138-0013-8000", "13800138000"),
    ],
)
def test_deterministic_field_normalization(field_type, raw_text, expected):
    field = StructuredFieldSpec(key="value", label="值", field_type=field_type)
    value, status, warnings = normalize_field_value(
        field=field,
        raw_text=raw_text,
        candidate_value=None,
    )
    assert value == expected
    assert status == "NORMALIZED"
    assert warnings == []


def test_zero_provider_records_become_explicit_review_items():
    service = object.__new__(StructuredExtractionService)
    service.settings = SimpleNamespace(
        structured_extraction_high_confidence=0.9,
        structured_extraction_retry_confidence=0.65,
        structured_extraction_max_retry_fields=8,
        structured_extraction_external_images_authorized=True,
    )
    service.extraction_provider = SimpleNamespace(
        supports_vision_retry=True,
        extract_with_image=lambda **_: None,
    )
    run = SimpleNamespace(document_id="doc-1", retry_strategy="INITIAL")
    field = StructuredFieldSpec(
        key="applicant",
        label="申请人",
        field_type="person_name",
        required=True,
    )
    result, rows = service._normalize_candidates(
        run=run,
        layout_extraction_run_id="layout-1",
        fields=[field],
        candidates=CandidateExtraction(records=[]),
        elements=[],
    )

    assert result.record_count == 1
    assert result.review_count == 1
    assert result.missing_required_field_count == 1
    assert result.retryable is False
    assert rows[0][1].status == "MISSING"


def test_llm_failure_falls_back_to_deterministic_key_value_mapping():
    """外部字段映射超时不能抹掉已成功取得的本地 OCR 证据。"""

    class FailingProvider:
        supports_vision_retry = False

        def extract(self, **_):
            raise LLMResponseError("timeout")

    service = object.__new__(StructuredExtractionService)
    service.settings = Settings(database_url="sqlite+pysqlite:///:memory:")
    service.agent_run_id = "agent-fallback"
    service.extraction_provider = FailingProvider()
    field = StructuredFieldSpec(
        key="applicant",
        label="申请人",
        field_type="person_name",
    )
    element = EvidenceElement(
        id="element-fallback",
        document_id="doc-fallback",
        extraction_run_id="layout-fallback",
        text="申请人：测试用户",
        page_number=1,
        bbox={"left": 10, "top": 20, "right": 200, "bottom": 60},
        metadata={"confidence": 0.96},
    )

    candidates = service._extract_candidates_with_fallback(
        run=SimpleNamespace(document_id="doc-fallback"),
        extraction_arguments={
            "fields": [field],
            "schema_mode": "EXPLICIT_FIELDS",
            "record_mode": "SINGLE_RECORD",
            "elements": [element],
            "max_records": 10,
        },
    )

    assert candidates.records[0].fields["applicant"].raw_text == "测试用户"
    assert candidates.records[0].fields["applicant"].evidence_element_ids == [
        "element-fallback"
    ]
    assert "LLM_FALLBACK_USED" in candidates.warnings


def test_deterministic_fallback_reconstructs_bbox_table_by_headers_and_row_anchors():
    """无单元格元数据时，仅按明确表头、序号锚点和列 bbox 重建记录。"""

    def element(element_id, text, left, top, right, bottom):
        return EvidenceElement(
            id=element_id,
            document_id="doc-table",
            extraction_run_id="layout-table",
            text=text,
            page_number=1,
            bbox={"left": left, "top": top, "right": right, "bottom": bottom},
            metadata={"confidence": 0.95},
        )

    elements = [
        element("h0", "序号", 80, 10, 140, 40),
        element("h1", "申请人", 170, 10, 250, 40),
        element("h2", "资助金额", 265, 10, 360, 40),
        element("h3", "使用情况", 380, 10, 500, 40),
        element("r1", "1", 115, 55, 140, 85),
        element("n1", "测试甲", 175, 50, 245, 90),
        element("a1", "1000", 275, 50, 345, 90),
        element("u1", "会议", 390, 50, 470, 90),
        element("d1", "2026.6.5", 510, 50, 590, 90),
        element("r2", "2", 115, 105, 140, 135),
        element("n2", "测试乙", 175, 100, 245, 140),
        element("a2", "2500", 275, 100, 345, 140),
        # 日期中心略越过序号中心中线，仍应按手写基线偏移归入第 2 行。
        element("d2", "2026.6.6", 510, 132, 590, 162),
        element("r3", "3", 115, 155, 140, 185),
        # 申请人签在右侧、金额和用途粘连时只生成低置信度真实 OCR 候选。
        element("a3", "3500.-（会议费）", 270, 150, 470, 190),
        element("n3", "测试丙", 520, 150, 590, 190),
    ]
    fields = [
        StructuredFieldSpec(key="applicant", label="申请人", field_type="person_name"),
        StructuredFieldSpec(key="amount", label="资助金额", field_type="money"),
        StructuredFieldSpec(key="date", label="申请日期", field_type="date"),
    ]

    result = DeterministicLayoutExtractionProvider().extract(
        fields=fields,
        schema_mode="EXPLICIT_FIELDS",
        record_mode="AUTO",
        elements=elements,
        max_records=10,
    )

    assert len(result.records) == 3
    assert result.records[0].fields["applicant"].raw_text == "测试甲"
    assert result.records[0].fields["amount"].raw_text == "1000"
    assert result.records[0].fields["date"].raw_text == "2026.6.5"
    assert result.records[1].fields["applicant"].evidence_element_ids == ["n2"]
    assert result.records[1].fields["date"].raw_text == "2026.6.6"
    assert result.records[2].fields["applicant"].raw_text == "测试丙"
    assert result.records[2].fields["applicant"].confidence <= 0.65
    assert result.records[2].fields["amount"].raw_text == "3500.-"
    assert result.records[2].fields["amount"].confidence <= 0.45


@pytest.mark.parametrize(
    ("raw_text", "expected"),
    [
        ("11500.-", {"amount": "11500", "currency": None}),
        ("3000-", {"amount": "3000", "currency": None}),
        ("1000.", {"amount": "1000", "currency": None}),
    ],
)
def test_money_normalization_accepts_handwritten_accounting_suffix(raw_text, expected):
    """完整数字后的手写记账尾缀不应让确定金额退化成待确认。"""

    value, status, warnings = normalize_field_value(
        field=StructuredFieldSpec(key="amount", label="金额", field_type="money"),
        raw_text=raw_text,
        candidate_value=raw_text,
    )

    assert value == expected
    assert status == "NORMALIZED"
    assert warnings == []


def test_low_confidence_field_enters_review_and_bounded_vision_retry():
    service = object.__new__(StructuredExtractionService)
    service.settings = SimpleNamespace(
        structured_extraction_high_confidence=0.9,
        structured_extraction_retry_confidence=0.65,
        structured_extraction_max_retry_fields=8,
        structured_extraction_external_images_authorized=True,
    )
    service.extraction_provider = SimpleNamespace(
        supports_vision_retry=True,
        extract_with_image=lambda **_: None,
    )
    run = SimpleNamespace(document_id="doc-1", retry_strategy="INITIAL")
    field = StructuredFieldSpec(
        key="amount",
        label="金额",
        field_type="money",
        required=True,
    )
    element = EvidenceElement(
        id="element-1",
        document_id="doc-1",
        extraction_run_id="layout-1",
        text="金额：10000元",
        page_number=1,
        bbox={"left": 10, "top": 20, "right": 100, "bottom": 60},
        metadata={},
    )
    candidates = CandidateExtraction.model_validate(
        {
            "records": [
                {
                    "record_index": 1,
                    "fields": {
                        "amount": {
                            "raw_text": "10000元",
                            "value": "10000元",
                            "confidence": 0.7,
                            "evidence_element_ids": ["element-1"],
                        }
                    },
                }
            ]
        }
    )

    result, rows = service._normalize_candidates(
        run=run,
        layout_extraction_run_id="layout-1",
        fields=[field],
        candidates=candidates,
        elements=[element],
    )

    assert rows[0][1].status == "NEEDS_REVIEW"
    assert "LOW_CONFIDENCE" in rows[0][1].warning_codes
    assert result.quality_band == "MEDIUM"
    assert result.retryable is True
    assert result.low_confidence_field_keys == ["amount"]


def test_evidence_backed_value_is_kept_and_unbacked_value_needs_review():
    service = object.__new__(StructuredExtractionService)
    service.settings = SimpleNamespace(
        structured_extraction_high_confidence=0.9,
        structured_extraction_retry_confidence=0.65,
        structured_extraction_max_retry_fields=8,
        structured_extraction_external_images_authorized=False,
    )
    service.extraction_provider = SimpleNamespace(supports_vision_retry=False)
    run = SimpleNamespace(document_id="doc-1", retry_strategy="INITIAL")
    fields = [StructuredFieldSpec(key="name", label="姓名", field_type="person_name")]
    candidates = CandidateExtraction.model_validate(
        {
            "records": [
                {
                    "record_index": 1,
                    "fields": {
                        "name": {
                            "raw_text": "金润逸",
                            "value": "金润逸",
                            "confidence": 0.96,
                            "evidence_element_ids": ["e1"],
                        }
                    },
                }
            ]
        }
    )
    element = EvidenceElement(
        id="e1",
        document_id="doc-1",
        extraction_run_id="layout-1",
        text="申请人 金润逸",
        page_number=1,
        bbox={"left": 1, "top": 2, "right": 3, "bottom": 4},
        metadata={},
    )
    result, _ = service._normalize_candidates(
        run=run,
        layout_extraction_run_id="layout-1",
        fields=fields,
        candidates=candidates,
        elements=[element],
    )
    assert result.records[0]["fields"]["name"]["status"] == "NORMALIZED"
    assert result.records[0]["fields"]["name"]["evidence"]["page_number"] == 1


def test_low_confidence_value_is_retryable_only_with_persisted_bbox_and_vision_permission():
    service = object.__new__(StructuredExtractionService)
    service.settings = SimpleNamespace(
        structured_extraction_high_confidence=0.9,
        structured_extraction_retry_confidence=0.65,
        structured_extraction_max_retry_fields=8,
        structured_extraction_external_images_authorized=True,
    )
    service.extraction_provider = SimpleNamespace(
        supports_vision_retry=True,
        extract_with_image=lambda **_: None,
    )
    candidates = CandidateExtraction.model_validate(
        {
            "records": [
                {
                    "record_index": 1,
                    "fields": {
                        "amount": {
                            "raw_text": "25xxx",
                            "value": None,
                            "confidence": 0.5,
                            "evidence_element_ids": ["e1"],
                        }
                    },
                }
            ]
        }
    )
    result, _ = service._normalize_candidates(
        run=SimpleNamespace(document_id="doc-1", retry_strategy="INITIAL"),
        layout_extraction_run_id="layout-1",
        fields=[StructuredFieldSpec(key="amount", label="金额", field_type="money")],
        candidates=candidates,
        elements=[
            EvidenceElement(
                id="e1",
                document_id="doc-1",
                extraction_run_id="layout-1",
                text="25xxx",
                page_number=1,
                bbox={"left": 1, "top": 2, "right": 30, "bottom": 12},
                metadata={},
            )
        ],
    )
    assert result.retryable is True
    assert result.low_confidence_field_keys == ["amount"]


def test_merge_only_replaces_target_with_higher_confidence_supported_value():
    initial = _tool_result(
        status="PARTIAL",
        quality_band="MEDIUM",
        field_status="NEEDS_REVIEW",
        confidence=0.5,
    )
    enhanced = _tool_result(
        status="COMPLETED",
        quality_band="HIGH",
        field_status="NORMALIZED",
        confidence=0.92,
    )
    enhanced["records"].append(
        {
            "record_index": 2,
            "fields": {
                "applicant": {
                    "normalized_value": "模型新增记录",
                    "confidence": 0.99,
                    "status": "NORMALIZED",
                }
            },
        }
    )
    merged = merge_structured_outputs(
        initial=initial,
        enhanced=enhanced,
        target_field_keys=["applicant"],
    )
    assert merged["records"][0]["fields"]["applicant"]["confidence"] == 0.92
    assert len(merged["records"]) == 1
    assert merged["retryable"] is False


def test_dispatch_budget_only_allows_observed_low_confidence_retry_fields():
    state = {
        "tool_invocations": [{"tool_name": "extract-image-structured-data"}],
        "observation": {
            "results": [
                {
                    "tool_name": "extract-image-structured-data",
                    "structured_extraction": {
                        "retryable": True,
                        "recommended_retry_strategy": "VISION_CROP",
                        "low_confidence_field_keys": ["amount"],
                    },
                }
            ]
        },
    }
    allowed = _structured_extraction_budget_error(
        tool_name="extract-image-structured-data",
        tool_input={"retry_strategy": "VISION_CROP", "target_field_keys": ["amount"]},
        state=state,
    )
    rejected = _structured_extraction_budget_error(
        tool_name="extract-image-structured-data",
        tool_input={"retry_strategy": "VISION_CROP", "target_field_keys": ["applicant"]},
        state=state,
    )
    assert allowed is None
    assert rejected is not None and rejected[0] == "STRUCTURED_EXTRACTION_RETRY_SCOPE_REJECTED"


def test_cached_parent_run_rebuilds_merged_child_result():
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    db = sessionmaker(bind=engine)()
    version = DocumentVersion(
        id="version-merge",
        document_id="doc-merge",
        storage_path="originals/form.png",
        filename="form.png",
        content_type="image/png",
        size_bytes=10,
        sha256="c" * 64,
    )
    schema = [
        {
            "key": "amount",
            "label": "金额",
            "field_type": "money",
            "required": True,
            "multiple": False,
            "aliases": [],
            "enum_values": [],
        }
    ]
    parent = _database_run(
        run_id="parent-run",
        document_id="doc-merge",
        version_id=version.id,
        schema=schema,
        retry_strategy="INITIAL",
    )
    child = _database_run(
        run_id="child-run",
        document_id="doc-merge",
        version_id=version.id,
        schema=schema,
        retry_strategy="VISION_CROP",
    )
    child.parent_run_id = parent.id
    child.target_field_keys_json = ["amount"]
    db.add_all(
        [
            version,
            parent,
            child,
            StructuredExtractionField(
                structured_extraction_run_id=parent.id,
                record_index=1,
                field_key="amount",
                field_label="金额",
                field_type="money",
                raw_text="25xxx",
                normalized_value_json=None,
                confidence=0.4,
                status="NEEDS_REVIEW",
                warning_codes_json=["NORMALIZATION_FAILED"],
            ),
            StructuredExtractionField(
                structured_extraction_run_id=child.id,
                record_index=1,
                field_key="amount",
                field_label="金额",
                field_type="money",
                raw_text="25000",
                normalized_value_json={"amount": "25000", "currency": None},
                confidence=0.93,
                status="NORMALIZED",
                warning_codes_json=[],
            ),
        ]
    )
    db.flush()
    service = object.__new__(StructuredExtractionService)
    service.db = db
    service.settings = SimpleNamespace(
        structured_extraction_max_retry_fields=8,
        structured_extraction_external_images_authorized=False,
    )
    service.extraction_provider = SimpleNamespace(supports_vision_retry=False)
    service.repository = StructuredExtractionRepository(db)

    result = service.result_for_run(parent)

    assert result.records[0]["fields"]["amount"]["confidence"] == 0.93
    assert result.review_count == 0
    assert result.missing_required_field_count == 0


def test_vision_retry_crop_uses_persisted_bbox_not_planner_coordinates(tmp_path: Path):
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    db = sessionmaker(bind=engine)()
    schema = [{"key": "amount", "label": "金额", "field_type": "money"}]
    parent = _database_run(
        run_id="crop-parent",
        document_id="crop-doc",
        version_id="crop-version",
        schema=schema,
        retry_strategy="INITIAL",
    )
    child = _database_run(
        run_id="crop-child",
        document_id="crop-doc",
        version_id="crop-version",
        schema=schema,
        retry_strategy="VISION_CROP",
    )
    child.parent_run_id = parent.id
    child.target_field_keys_json = ["amount"]
    layout_run = DocumentExtractionRun(
        id="crop-layout",
        document_id="crop-doc",
        document_version_id="crop-version",
        status="COMPLETED",
        extractor="fake",
    )
    db.add_all(
        [
            parent,
            child,
            layout_run,
            StructuredExtractionField(
                structured_extraction_run_id=parent.id,
                record_index=1,
                field_key="amount",
                field_label="金额",
                field_type="money",
                raw_text="25xxx",
                confidence=0.4,
                status="NEEDS_REVIEW",
                page_number=1,
                bbox_json={"left": 100, "top": 120, "right": 240, "bottom": 180},
            ),
        ]
    )
    db.flush()
    image_path = tmp_path / "source.png"
    Image.new("RGB", (1000, 800), "white").save(image_path)
    service = object.__new__(StructuredExtractionService)
    service.db = db

    data_url = service._targeted_crop_data_url(
        run=child,
        source_path=image_path,
        content_type="image/png",
    )
    crop = Image.open(io.BytesIO(base64.b64decode(data_url.split(",", 1)[1])))

    assert data_url.startswith("data:image/png;base64,")
    assert crop.width < 1000 and crop.height < 800
    assert crop.width == 376 and crop.height == 216

    candidates = CandidateExtraction.model_validate(
        {
            "records": [
                {
                    "record_index": 1,
                    "fields": {
                        "amount": {
                            "raw_text": "25000",
                            "value": "25000",
                            "confidence": 0.94,
                            "evidence_element_ids": [],
                        }
                    },
                }
            ]
        }
    )
    evidence = StructuredExtractionRepository(db).append_vision_candidate_evidence(
        run=child,
        layout_extraction_run_id=layout_run.id,
        candidates=candidates,
    )
    assert evidence[0].text == "25000"
    assert evidence[0].bbox == {"left": 100, "top": 120, "right": 240, "bottom": 180}
    assert candidates.records[0].fields["amount"].evidence_element_ids == [evidence[0].id]
    assert db.query(DocumentElement).filter(DocumentElement.label == "vision_crop_text").count() == 1


def test_local_vlm_can_retry_missing_field_on_single_page_and_persist_its_bbox():
    """没有基础 OCR bbox 的缺失字段只能由单页本地 VLM 建立新证据。"""

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    db = sessionmaker(bind=engine)()
    schema = [{"key": "date", "label": "申请日期", "field_type": "date"}]
    parent = _database_run(
        run_id="vl-parent",
        document_id="vl-doc",
        version_id="vl-version",
        schema=schema,
        retry_strategy="INITIAL",
    )
    child = _database_run(
        run_id="vl-child",
        document_id="vl-doc",
        version_id="vl-version",
        schema=schema,
        retry_strategy="VISION_CROP",
    )
    child.parent_run_id = parent.id
    child.target_field_keys_json = ["date"]
    layout_run = DocumentExtractionRun(
        id="vl-layout",
        document_id="vl-doc",
        document_version_id="vl-version",
        status="COMPLETED",
        extractor="fake",
    )
    parent.layout_extraction_run_id = layout_run.id
    db.add_all(
        [
            parent,
            child,
            layout_run,
            DocumentPage(
                document_id="vl-doc",
                extraction_run_id=layout_run.id,
                page_number=1,
                text_content="",
                metadata_json={"width": 1000, "height": 800},
            ),
            StructuredExtractionField(
                structured_extraction_run_id=parent.id,
                record_index=1,
                field_key="date",
                field_label="申请日期",
                field_type="date",
                raw_text=None,
                confidence=0,
                status="MISSING",
                bbox_json={},
            ),
        ]
    )
    db.flush()
    service = object.__new__(StructuredExtractionService)
    service.db = db
    service.settings = SimpleNamespace(
        structured_extraction_high_confidence=0.9,
        structured_extraction_retry_confidence=0.65,
        structured_extraction_max_retry_fields=8,
        structured_extraction_external_images_authorized=False,
    )
    service.extraction_provider = SimpleNamespace(supports_vision_retry=False)
    service.vision_provider = SimpleNamespace(
        enabled=True,
        is_external=False,
        supports_unlocated_retry=True,
        name="paddleocr_vl",
    )

    initial, _ = service._normalize_candidates(
        run=parent,
        layout_extraction_run_id=layout_run.id,
        fields=[StructuredFieldSpec(key="date", label="申请日期", field_type="date")],
        candidates=CandidateExtraction(records=[]),
        elements=[],
    )
    assert initial.retryable is True
    assert initial.low_confidence_field_keys == ["date"]

    candidates = CandidateExtraction.model_validate(
        {
            "records": [
                {
                    "record_index": 1,
                    "fields": {
                        "date": {
                            "raw_text": "2026.6.5",
                            "value": "2026.6.5",
                            "confidence": 0.92,
                            "evidence_element_ids": ["vision-transient:0"],
                        }
                    },
                }
            ]
        }
    )
    evidence = StructuredExtractionRepository(db).append_vision_candidate_evidence(
        run=child,
        layout_extraction_run_id=layout_run.id,
        candidates=candidates,
        vision_elements=[
            EvidenceElement(
                id="vision-transient:0",
                document_id="vl-doc",
                extraction_run_id=layout_run.id,
                text="申请日期 2026.6.5",
                page_number=1,
                bbox={"left": 300, "top": 120, "right": 460, "bottom": 160},
                metadata={"source": "paddleocr_vl"},
            )
        ],
    )
    assert evidence[0].page_number == 1
    assert evidence[0].bbox == {"left": 300.0, "top": 120.0, "right": 460.0, "bottom": 160.0}
    assert candidates.records[0].fields["date"].evidence_element_ids == [evidence[0].id]


@pytest.mark.parametrize("presentation", ["CSV", "XLSX"])
def test_csv_and_xlsx_export_create_persisted_derivative(tmp_path: Path, presentation: str):
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    db = sessionmaker(bind=engine)()
    version = DocumentVersion(
        id="version-1",
        document_id="doc-1",
        storage_path="originals/form.png",
        filename="form.png",
        content_type="image/png",
        size_bytes=10,
        sha256="a" * 64,
    )
    run = StructuredExtractionRun(
        id="run-1",
        document_id="doc-1",
        document_version_id=version.id,
        schema_mode="EXPLICIT_FIELDS",
        field_schema_json=[],
        schema_fingerprint="b" * 64,
        record_mode="TABLE_ROWS",
        presentation=presentation,
        provider="fake",
        model_name="fake",
        prompt_version="v1",
        retry_strategy="INITIAL",
        status="COMPLETED",
    )
    db.add_all([version, run])
    db.flush()
    result = StructuredExtractionResult(
        field_schema=[{"key": "name", "label": "姓名", "field_type": "person_name"}],
        records=[
            {
                "record_index": 1,
                "fields": {
                    "name": {
                        "normalized_value": "金润逸",
                    }
                },
            }
        ],
        review_items=[],
        record_count=1,
        field_count=1,
        review_count=0,
        missing_required_field_count=0,
        quality_score=0.96,
        quality_band="HIGH",
        retryable=False,
        recommended_retry_strategy="NONE",
        low_confidence_field_keys=[],
    )
    artifact = StructuredExtractionExportService(
        db=db,
        settings=Settings(database_url="postgresql://test", file_storage_root=str(tmp_path)),
    ).ensure_export(run=run, result=result)
    path = tmp_path / "derivatives" / "structured-extraction" / "doc-1" / f"run-1.{presentation.lower()}"

    assert artifact is not None and artifact["format"] == presentation
    assert path.is_file()
    if presentation == "CSV":
        with path.open(encoding="utf-8-sig", newline="") as handle:
            assert list(csv.reader(handle)) == [["姓名"], ["金润逸"]]
    else:
        workbook = load_workbook(path, read_only=True)
        assert list(workbook.active.values) == [("姓名",), ("金润逸",)]


@pytest.mark.parametrize("presentation", ["CSV", "XLSX"])
def test_structured_export_escapes_spreadsheet_formulas(tmp_path: Path, presentation: str):
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    db = sessionmaker(bind=engine)()
    version = DocumentVersion(
        id="version-formula",
        document_id="doc-formula",
        storage_path="originals/form.png",
        filename="form.png",
        content_type="image/png",
        size_bytes=10,
        sha256="f" * 64,
    )
    run = StructuredExtractionRun(
        id="run-formula",
        document_id="doc-formula",
        document_version_id=version.id,
        schema_mode="EXPLICIT_FIELDS",
        field_schema_json=[],
        schema_fingerprint="e" * 64,
        record_mode="TABLE_ROWS",
        presentation=presentation,
        provider="fake",
        model_name="fake",
        prompt_version="v1",
        retry_strategy="INITIAL",
        status="COMPLETED",
    )
    db.add_all([version, run])
    db.flush()
    result = StructuredExtractionResult(
        field_schema=[{"key": "value", "label": "=危险表头", "field_type": "string"}],
        records=[
            {
                "record_index": 1,
                "fields": {"value": {"normalized_value": "+SUM(1,1)"}},
            }
        ],
        review_items=[],
        record_count=1,
        field_count=1,
        review_count=0,
        missing_required_field_count=0,
        quality_score=0.96,
        quality_band="HIGH",
        retryable=False,
        recommended_retry_strategy="NONE",
        low_confidence_field_keys=[],
    )
    StructuredExtractionExportService(
        db=db,
        settings=Settings(database_url="postgresql://test", file_storage_root=str(tmp_path)),
    ).ensure_export(run=run, result=result)
    path = (
        tmp_path
        / "derivatives"
        / "structured-extraction"
        / "doc-formula"
        / f"run-formula.{presentation.lower()}"
    )

    if presentation == "CSV":
        with path.open(encoding="utf-8-sig", newline="") as handle:
            assert list(csv.reader(handle)) == [["'=危险表头"], ["'+SUM(1,1)"]]
    else:
        workbook = load_workbook(path, read_only=True, data_only=False)
        cells = list(workbook.active.iter_rows())[0:2]
        assert cells[0][0].value == "'=危险表头"
        assert cells[1][0].value == "'+SUM(1,1)"
        assert cells[1][0].data_type != "f"


def test_async_failure_persists_run_and_failure_changeset():
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    db = sessionmaker(bind=engine)()
    agent_run = AgentRun(
        id="agent-failed",
        conversation_id="conversation-failed",
        message_id="message-failed",
        user_id="user-failed",
        status="WAITING_FOR_ASYNC_JOB",
        graph_state_json={"status": "WAITING_FOR_ASYNC_JOB", "async_job_ids": ["job-failed"]},
    )
    structured_run = StructuredExtractionRun(
        id="structured-failed",
        document_id="doc-failed",
        document_version_id="version-failed",
        agent_run_id=agent_run.id,
        schema_mode="EXPLICIT_FIELDS",
        field_schema_json=[],
        schema_fingerprint="d" * 64,
        record_mode="SINGLE_RECORD",
        presentation="JSON",
        provider="fake",
        model_name="fake",
        prompt_version="v1",
        retry_strategy="INITIAL",
        status="FAILED",
        error_code="OCR_PROVIDER_TEMPORARY_FAILURE",
        error_message="腾讯云 OCR 暂时不可用，请稍后重试。",
    )
    invocation = ToolInvocation(
        id="invocation-failed",
        agent_run_id=agent_run.id,
        tool_name="extract-image-structured-data",
        input_json={},
        output_json={"kind": "filesystem_job", "status": "WAITING_FOR_ASYNC_JOB"},
        status="WAITING_FOR_ASYNC_JOB",
    )
    job = FilesystemJob(
        id="job-failed",
        job_type="STRUCTURED_IMAGE_EXTRACTION",
        queue_name="STRUCTURED_EXTRACTION",
        status="FAILED",
        payload_json={
            "structured_extraction_run_id": structured_run.id,
            "agent_run_id": agent_run.id,
        },
    )
    db.add_all([agent_run, structured_run, invocation, job])
    db.flush()

    assert fail_structured_extraction_agent_run(
        db=db,
        job=job,
        error_message="图片结构化抽取失败。",
    ) is True
    db.flush()

    assert structured_run.status == "FAILED"
    assert structured_run.error_code == "OCR_PROVIDER_TEMPORARY_FAILURE"
    assert agent_run.status == "FAILED"
    assert invocation.status == "FAILED"
    assert invocation.output_json["error"]["code"] == "OCR_PROVIDER_TEMPORARY_FAILURE"
    assert agent_run.changeset_id == invocation.changeset_id
    failure = db.query(ChangeItem).filter(
        ChangeItem.change_type == "STRUCTURED_EXTRACTION_FAILED"
    ).one()
    assert failure.after_value_json["original_unchanged"] is True
    assert failure.after_value_json["error_code"] == "OCR_PROVIDER_TEMPORARY_FAILURE"

    structured_run.status = "COMPLETED"
    _resume_agent_run(
        db=db,
        job=job,
        run=structured_run,
        output={
            "kind": "structured_image_extraction",
            "ok": True,
            "status": "COMPLETED",
            "document_id": structured_run.document_id,
            "structured_extraction_run_id": structured_run.id,
            "record_count": 1,
            "field_count": 2,
            "review_count": 0,
            "quality_band": "HIGH",
            "original_unchanged": True,
        },
        clarification=None,
    )

    assert agent_run.status == "COMPLETED"
    assert agent_run.error_message is None
    assert agent_run.graph_state_json["errors"] == []


def test_structured_worker_public_error_does_not_expose_exception_details():
    job = FilesystemJob(
        job_type="STRUCTURED_IMAGE_EXTRACTION",
        queue_name="STRUCTURED_EXTRACTION",
    )

    message = _public_job_error_message(
        job=job,
        error=RuntimeError("/secret/source.png api-key-value"),
    )

    assert "/secret/source.png" not in message
    assert "api-key-value" not in message


def test_user_receipt_projects_dynamic_result_and_export_without_internal_paths():
    output = _tool_result(
        status="COMPLETED",
        quality_band="HIGH",
        field_status="NORMALIZED",
        confidence=0.96,
    )
    output["export_artifact"] = {
        "artifact_id": "artifact-1",
        "format": "XLSX",
        "filename": "result.xlsx",
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "size_bytes": 123,
        "storage_path": "/secret/path/result.xlsx",
    }
    result = AgentRunResult(
        agent_run_id="run-1",
        conversation_id="conv-1",
        user_id="user-1",
        message_id="message-1",
        intent="EXTRACT_STRUCTURED_DATA",
        status="COMPLETED",
        selected_skills=["image-structured-extraction"],
        tool_plan={},
        tool_results=[],
        tool_invocations=[
            ToolInvocationRecord(
                tool_name="extract-image-structured-data",
                input_json={},
                output_json=output,
                status="COMPLETED",
            )
        ],
        final_response="已完成",
    )
    receipt = build_user_task_receipt(result)
    projected = receipt.structured_extraction_result

    assert receipt.response_type == "structured_extraction"
    assert projected is not None
    assert projected["records"][0]["fields"]["applicant"]["normalized_value"] == "金润逸"
    assert projected["export_artifact"]["artifact_id"] == "artifact-1"
    assert "storage_path" not in projected["export_artifact"]


def test_user_receipt_preserves_custom_column_count_and_order():
    """回执投影必须完整保留后端验证后的动态显示列。"""

    output = _tool_result(
        status="COMPLETED",
        quality_band="HIGH",
        field_status="NORMALIZED",
        confidence=0.96,
    )
    schema = [
        {"key": "applicant", "label": "申请人", "field_type": "person_name"},
        {"key": "amount", "label": "资助金额", "field_type": "money"},
        {"key": "date", "label": "申请日期", "field_type": "date"},
        {"key": "usage", "label": "使用情况摘要", "field_type": "string"},
        {"key": "remark", "label": "备注", "field_type": "string"},
    ]
    output["presentation"] = "TABLE"
    output["field_schema"] = schema
    output["field_count"] = len(schema)
    output["records"] = [
        {
            "record_index": 1,
            "fields": {
                item["key"]: {
                    "raw_text": item["label"],
                    "normalized_value": item["label"],
                    "confidence": 0.96,
                    "status": "NORMALIZED",
                    "evidence": {"page_number": 1, "bbox": {}},
                    "warnings": [],
                }
                for item in schema
            },
        }
    ]
    result = AgentRunResult(
        agent_run_id="run-columns",
        conversation_id="conv-columns",
        user_id="user-1",
        message_id="message-columns",
        intent="EXTRACT_STRUCTURED_DATA",
        status="COMPLETED",
        selected_skills=["image-structured-extraction"],
        tool_plan={},
        tool_results=[],
        tool_invocations=[
            ToolInvocationRecord(
                tool_name="extract-image-structured-data",
                input_json={},
                output_json=output,
                status="COMPLETED",
            )
        ],
        final_response="已完成",
    )

    projected = build_user_task_receipt(result).structured_extraction_result

    assert projected is not None
    assert [item["label"] for item in projected["field_schema"]] == [
        "申请人",
        "资助金额",
        "申请日期",
        "使用情况摘要",
        "备注",
    ]
    assert len(projected["field_schema"]) == 5
    assert list(projected["records"][0]["fields"]) == [
        "applicant",
        "amount",
        "date",
        "usage",
        "remark",
    ]


def test_graph_summary_does_not_copy_structured_field_values():
    output = _tool_result(
        status="COMPLETED",
        quality_band="HIGH",
        field_status="NORMALIZED",
        confidence=0.96,
    )

    summary = _structured_graph_summary(output)

    assert summary["record_count"] == 1
    assert summary["structured_extraction_run_id"] == "structured-run-1"
    assert "records" not in summary
    assert "field_schema" not in summary
    assert "review_items" not in summary


def _tool_result(*, status: str, quality_band: str, field_status: str, confidence: float):
    return {
        "kind": "structured_image_extraction",
        "ok": True,
        "status": status,
        "document_id": "doc-1",
        "structured_extraction_run_id": "structured-run-1",
        "schema_mode": "EXPLICIT_FIELDS",
        "record_mode": "SINGLE_RECORD",
        "presentation": "XLSX",
        "field_schema": [
            {"key": "applicant", "label": "申请人", "field_type": "person_name", "required": True}
        ],
        "records": [
            {
                "record_index": 1,
                "fields": {
                    "applicant": {
                        "raw_text": "金润逸",
                        "normalized_value": "金润逸",
                        "confidence": confidence,
                        "status": field_status,
                        "evidence": {
                            "page_number": 1,
                            "bbox": {"left": 1, "top": 2, "right": 3, "bottom": 4},
                        },
                        "warnings": [],
                    }
                },
            }
        ],
        "review_items": [],
        "record_count": 1,
        "field_count": 1,
        "review_count": 0 if field_status == "NORMALIZED" else 1,
        "missing_required_field_count": 0,
        "quality_score": confidence,
        "quality_band": quality_band,
        "retryable": field_status != "NORMALIZED",
        "recommended_retry_strategy": "NONE" if field_status == "NORMALIZED" else "VISION_CROP",
        "low_confidence_field_keys": [] if field_status == "NORMALIZED" else ["applicant"],
        "original_unchanged": True,
    }


def _database_run(
    *,
    run_id: str,
    document_id: str,
    version_id: str,
    schema: list[dict],
    retry_strategy: str,
) -> StructuredExtractionRun:
    return StructuredExtractionRun(
        id=run_id,
        document_id=document_id,
        document_version_id=version_id,
        schema_mode="EXPLICIT_FIELDS",
        field_schema_json=schema,
        schema_fingerprint=f"fingerprint-{run_id}",
        record_mode="SINGLE_RECORD",
        presentation="JSON",
        provider="fake",
        model_name="fake",
        prompt_version="v1",
        retry_strategy=retry_strategy,
        status="COMPLETED" if retry_strategy == "VISION_CROP" else "PARTIAL",
        record_count=1,
        review_count=0 if retry_strategy == "VISION_CROP" else 1,
        quality_score=0.93 if retry_strategy == "VISION_CROP" else 0.4,
        quality_band="HIGH" if retry_strategy == "VISION_CROP" else "MEDIUM",
    )
